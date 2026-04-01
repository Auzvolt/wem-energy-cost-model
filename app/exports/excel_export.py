"""Excel workbook export for scenario results — issue #43.

Generates a multi-sheet .xlsx workbook from simulation/financial results using
openpyxl.  Sheets produced:

- **Summary** — key financial metrics (NPV, IRR, LCOE, payback)
- **Cashflow** — year-by-year cashflow table from ``results["cashflow"]``
- **Dispatch** — dispatch profile (optional; only written when ``results["dispatch_profile"]``
  is a non-empty DataFrame)
"""

from __future__ import annotations

import logging
import os
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

# Visible column order / friendly headers for the cashflow sheet
_CASHFLOW_COLUMNS: list[tuple[str, str]] = [
    ("year", "Year"),
    ("energy_revenue", "Energy Revenue ($)"),
    ("fcess_revenue", "FCESS Revenue ($)"),
    ("capacity_revenue", "Capacity Revenue ($)"),
    ("network_savings", "Network Savings ($)"),
    ("total_revenue", "Total Revenue ($)"),
    ("opex_fixed", "Fixed OPEX ($)"),
    ("opex_variable", "Variable OPEX ($)"),
    ("opex_total", "Total OPEX ($)"),
    ("replacement_capex", "Replacement CAPEX ($)"),
    ("debt_service", "Debt Service ($)"),
    ("ebitda", "EBITDA ($)"),
    ("fcff", "FCFF ($)"),
    ("fcfe", "FCFE ($)"),
    ("fcfe_discounted", "Discounted FCFE ($)"),
]

_DISPATCH_COLUMNS: list[tuple[str, str]] = [
    ("interval", "Interval"),
    ("dispatch_kw", "Dispatch (kW)"),
    ("soc_pct", "State of Charge (%)"),
]


class ExcelExporter:
    """Generates multi-sheet Excel workbooks from simulation results."""

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export(self, results: dict[str, Any], output_path: str) -> str:
        """Generate an Excel workbook and write it to *output_path*.

        Args:
            results: Dictionary that **must** contain:
                - ``"scenario_name"`` (str) — used in the Summary sheet title.
                - ``"cashflow"`` (pd.DataFrame | None) — year-by-year cashflow.
                - ``"financial_summary"`` (dict) — keys: npv, irr, lcoe, payback_years.
                - ``"dispatch_profile"`` (pd.DataFrame | None) — optional dispatch data.
            output_path: Destination file path (will be created / overwritten).

        Returns:
            Absolute path to the generated ``.xlsx`` file.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Remove the default empty sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        # ---- Summary sheet ----
        ws_summary = wb.create_sheet("Summary")
        self._write_summary(
            ws_summary, results, Font=Font, PatternFill=PatternFill, Alignment=Alignment
        )

        # ---- Cashflow sheet ----
        cashflow: pd.DataFrame | None = results.get("cashflow")
        ws_cf = wb.create_sheet("Cashflow")
        self._write_cashflow(
            ws_cf, cashflow, Font=Font, PatternFill=PatternFill, get_column_letter=get_column_letter
        )

        # ---- Dispatch sheet (optional) ----
        dispatch: pd.DataFrame | None = results.get("dispatch_profile")
        if dispatch is not None and not dispatch.empty:
            ws_dp = wb.create_sheet("Dispatch")
            self._write_dispatch(
                ws_dp,
                dispatch,
                Font=Font,
                PatternFill=PatternFill,
                get_column_letter=get_column_letter,
            )

        abs_path = os.path.abspath(output_path)
        wb.save(abs_path)
        logger.info("Excel workbook written to %s", abs_path)
        return abs_path

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _header_fill(self, PatternFill: Any) -> Any:
        return PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

    def _header_font(self, Font: Any) -> Any:
        return Font(bold=True, color="FFFFFF", size=11)

    def _write_summary(
        self,
        ws: Any,
        results: dict[str, Any],
        *,
        Font: Any,
        PatternFill: Any,
        Alignment: Any,
    ) -> None:
        """Write the Summary sheet."""
        scenario_name: str = str(results.get("scenario_name", "Scenario"))
        ws.title = "Summary"

        title_font = Font(bold=True, size=14)
        ws["A1"] = f"Scenario Report: {scenario_name}"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")
        ws.row_dimensions[1].height = 24

        ws.append([])  # blank row

        header_font = self._header_font(Font)
        header_fill = self._header_fill(PatternFill)

        ws.append(["Metric", "Value"])
        header_row = ws.max_row
        for col in (1, 2):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        summary: dict[str, Any] = results.get("financial_summary") or {}
        rows: list[tuple[str, Any]] = [
            ("NPV ($)", summary.get("npv", "N/A")),
            ("IRR", f"{summary['irr']:.1%}" if "irr" in summary else "N/A"),
            ("LCOE ($/MWh)", summary.get("lcoe", "N/A")),
            ("Payback Period (years)", summary.get("payback_years", "N/A")),
        ]
        for label, value in rows:
            ws.append([label, value])

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

    def _write_cashflow(
        self,
        ws: Any,
        df: pd.DataFrame | None,
        *,
        Font: Any,
        PatternFill: Any,
        get_column_letter: Any,
    ) -> None:
        """Write the Cashflow sheet."""
        header_font = self._header_font(Font)
        header_fill = self._header_fill(PatternFill)

        if df is None or df.empty:
            ws.append(["No cashflow data available."])
            return

        # Build ordered column list — keep only columns that exist in the df
        present_cols = [(col, label) for col, label in _CASHFLOW_COLUMNS if col in df.columns]
        # Append any extra columns not in our defined list
        known = {c for c, _ in _CASHFLOW_COLUMNS}
        for col in df.columns:
            if col not in known:
                present_cols.append((col, col))

        headers = [label for _, label in present_cols]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for _, row in df.iterrows():
            ws.append([row.get(col) for col, _ in present_cols])

        # Auto-width
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

    def _write_dispatch(
        self,
        ws: Any,
        df: pd.DataFrame,
        *,
        Font: Any,
        PatternFill: Any,
        get_column_letter: Any,
    ) -> None:
        """Write the Dispatch sheet."""
        header_font = self._header_font(Font)
        header_fill = self._header_fill(PatternFill)

        present_cols = [(col, label) for col, label in _DISPATCH_COLUMNS if col in df.columns]
        known = {c for c, _ in _DISPATCH_COLUMNS}
        for col in df.columns:
            if col not in known:
                present_cols.append((col, col))

        headers = [label for _, label in present_cols]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for _, row in df.iterrows():
            ws.append([row.get(col) for col, _ in present_cols])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
