"""Assumption library import/export — JSON and Excel formats.

Provides serialisation and deserialisation of AssumptionSet objects so that
assumption packages can be transported between environments (dev → staging →
production) or shared with external stakeholders as Excel workbooks.

Supported formats:
- JSON  : lossless round-trip; preserves all fields including UUIDs and dates.
- Excel : human-readable workbook with one sheet per category; suitable for
          review and bulk editing before re-import.
"""

from __future__ import annotations

import contextlib
import io
import json
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app.assumptions.models import AssumptionCategory, AssumptionEntry, AssumptionSet

# ---------------------------------------------------------------------------
# JSON export / import
# ---------------------------------------------------------------------------


def _json_default(obj: Any) -> Any:
    """Custom JSON encoder for types not natively serialisable."""
    if isinstance(obj, (uuid.UUID, date, datetime)):
        return str(obj)
    raise TypeError(f"Object of type {type(obj)} is not JSON serialisable")


def export_json(assumption_set: AssumptionSet) -> str:
    """Serialise an AssumptionSet to a JSON string.

    The output is a canonical JSON representation that can be round-tripped
    through :func:`import_json` without loss of information.

    Args:
        assumption_set: The set to export.

    Returns:
        A pretty-printed JSON string.
    """
    data = assumption_set.model_dump(mode="json")
    return json.dumps(data, indent=2, default=_json_default)


def export_json_file(assumption_set: AssumptionSet, path: Path | str) -> None:
    """Write an AssumptionSet as JSON to *path*.

    Args:
        assumption_set: The set to export.
        path: Destination file path (will be overwritten if it exists).
    """
    Path(path).write_text(export_json(assumption_set), encoding="utf-8")


def import_json(data: str | bytes | dict[str, Any]) -> AssumptionSet:
    """Deserialise an AssumptionSet from a JSON string, bytes, or dict.

    Args:
        data: JSON text, bytes, or a pre-parsed dict.

    Returns:
        A validated AssumptionSet instance.

    Raises:
        ValueError: If the data is missing required fields or cannot be
            coerced into a valid AssumptionSet.
        json.JSONDecodeError: If *data* is a string/bytes with invalid JSON.
    """
    if isinstance(data, (str, bytes)):
        parsed: dict[str, Any] = json.loads(data)
    else:
        parsed = data

    return AssumptionSet.model_validate(parsed)


def import_json_file(path: Path | str) -> AssumptionSet:
    """Load an AssumptionSet from a JSON file.

    Args:
        path: Path to the JSON file produced by :func:`export_json_file`.

    Returns:
        A validated AssumptionSet instance.
    """
    raw = Path(path).read_text(encoding="utf-8")
    return import_json(raw)


# ---------------------------------------------------------------------------
# Excel export / import
# ---------------------------------------------------------------------------

# Column order for the entries sheet.
_ENTRY_COLUMNS = [
    "id",
    "set_id",
    "category",
    "key",
    "value",
    "unit",
    "source",
    "created_at",
]

# Metadata sheet field order.
_META_FIELDS = [
    "id",
    "name",
    "description",
    "author",
    "created_at",
    "effective_from",
    "superseded_by",
]


def export_excel(assumption_set: AssumptionSet) -> bytes:
    """Serialise an AssumptionSet to an Excel workbook as raw bytes.

    The workbook contains:
    - A **Metadata** sheet with the set-level fields.
    - One sheet per :class:`~app.assumptions.models.AssumptionCategory` containing
      the entries belonging to that category.
    - An **All Entries** sheet with every entry across all categories.

    The ``value`` column contains the JSON-encoded representation of each
    entry's value so that complex structures (dicts, lists) survive the
    round-trip through :func:`import_excel`.

    Args:
        assumption_set: The set to export.

    Returns:
        Raw bytes of an ``.xlsx`` workbook.

    Raises:
        ImportError: If ``openpyxl`` is not installed.
    """
    try:
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()

    # ---- Metadata sheet ------------------------------------------------
    ws_meta = wb.active
    assert ws_meta is not None
    ws_meta.title = "Metadata"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws_meta.append(["Field", "Value"])
    for cell in ws_meta[1]:
        cell.font = header_font
        cell.fill = header_fill

    set_dict = assumption_set.model_dump(mode="json")
    for field in _META_FIELDS:
        value = set_dict.get(field)
        ws_meta.append([field, str(value) if value is not None else ""])

    # ---- Helper: write entries to a worksheet --------------------------
    def _write_entries_sheet(ws: Any, entries: list[AssumptionEntry]) -> None:
        ws.append(_ENTRY_COLUMNS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for entry in entries:
            ed = entry.model_dump(mode="json")
            ws.append(
                [
                    str(ed["id"]),
                    str(ed["set_id"]),
                    str(ed["category"]),
                    ed["key"],
                    json.dumps(ed["value"]),  # encode complex values
                    ed.get("unit") or "",
                    ed.get("source") or "",
                    str(ed.get("created_at") or ""),
                ]
            )

    # ---- All Entries sheet ---------------------------------------------
    ws_all = wb.create_sheet("All Entries")
    _write_entries_sheet(ws_all, assumption_set.entries)

    # ---- Per-category sheets -------------------------------------------
    categories_present = {e.category for e in assumption_set.entries}
    for cat in AssumptionCategory:
        if cat not in categories_present:
            continue
        cat_entries = [e for e in assumption_set.entries if e.category == cat]
        ws_cat = wb.create_sheet(cat.value.replace("_", " ").title())
        _write_entries_sheet(ws_cat, cat_entries)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_excel_file(assumption_set: AssumptionSet, path: Path | str) -> None:
    """Write an AssumptionSet as an Excel workbook to *path*.

    Args:
        assumption_set: The set to export.
        path: Destination file path (will be overwritten if it exists).
    """
    Path(path).write_bytes(export_excel(assumption_set))


def import_excel(data: bytes) -> AssumptionSet:
    """Deserialise an AssumptionSet from an Excel workbook byte string.

    Reads the **Metadata** sheet for set-level fields and the **All Entries**
    sheet for entry data. Per-category sheets are ignored on import to avoid
    duplication.

    Args:
        data: Raw bytes of an ``.xlsx`` workbook produced by :func:`export_excel`.

    Returns:
        A validated AssumptionSet instance.

    Raises:
        ImportError: If ``openpyxl`` is not installed.
        ValueError: If the workbook is missing required sheets or columns.
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for Excel import. Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # ---- Read metadata -------------------------------------------------
    if "Metadata" not in wb.sheetnames:
        raise ValueError("Workbook is missing the 'Metadata' sheet")

    ws_meta = wb["Metadata"]
    meta: dict[str, Any] = {}
    for row in ws_meta.iter_rows(min_row=2, values_only=True):
        field, value = row[0], row[1]
        if field:
            meta[str(field)] = value if value != "" else None

    # ---- Read entries --------------------------------------------------
    if "All Entries" not in wb.sheetnames:
        raise ValueError("Workbook is missing the 'All Entries' sheet")

    ws_all = wb["All Entries"]
    rows = list(ws_all.iter_rows(values_only=True))
    if not rows:
        entries_data: list[dict[str, Any]] = []
    else:
        header = [str(h) for h in rows[0]]
        entries_data = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = dict(zip(header, row, strict=False))
            # Decode the JSON-encoded value column
            raw_value = row_dict.get("value")
            if raw_value and isinstance(raw_value, str):
                with contextlib.suppress(json.JSONDecodeError):
                    row_dict["value"] = json.loads(raw_value)
            entries_data.append(row_dict)

    meta["entries"] = entries_data
    return AssumptionSet.model_validate(meta)


def import_excel_file(path: Path | str) -> AssumptionSet:
    """Load an AssumptionSet from an Excel workbook file.

    Args:
        path: Path to the ``.xlsx`` workbook produced by :func:`export_excel_file`.

    Returns:
        A validated AssumptionSet instance.
    """
    raw = Path(path).read_bytes()
    return import_excel(raw)
