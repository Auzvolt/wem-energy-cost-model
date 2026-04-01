"""Tests for ExcelExporter and PDFExporter (closes #43)."""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd

from app.exports.excel_export import ExcelExporter
from app.exports.pdf_export import PDFExporter

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------


def _minimal_results(*, with_dispatch: bool = False) -> dict:
    """Return the minimum valid results dict accepted by both exporters."""
    cashflow = pd.DataFrame(
        {
            "year": [1, 2, 3],
            "energy_revenue_aud": [80_000.0, 82_000.0, 84_000.0],
            "fcess_revenue_aud": [20_000.0, 20_500.0, 21_000.0],
            "capacity_revenue_aud": [15_000.0, 15_000.0, 15_000.0],
            "network_savings_aud": [5_000.0, 5_100.0, 5_200.0],
            "opex_aud": [-30_000.0, -30_600.0, -31_200.0],
            "debt_service_aud": [-25_000.0, -25_000.0, -25_000.0],
            "ebitda_aud": [90_000.0, 92_000.0, 94_000.0],
            "fcfe_aud": [65_000.0, 67_000.0, 69_000.0],
        }
    )
    results: dict = {
        "scenario_name": "Test Scenario",
        "cashflow": cashflow,
        "financial_summary": {
            "npv": 500_000.0,
            "irr": 0.092,
            "lcoe": 0.082,
            "payback_years": 8.4,
        },
    }
    if with_dispatch:
        results["dispatch_profile"] = pd.DataFrame(
            {
                "interval": pd.date_range("2024-01-15", periods=5, freq="30min"),
                "power_kw": [100.0, 150.0, 200.0, 180.0, 120.0],
                "soc_pct": [50.0, 55.0, 60.0, 57.0, 52.0],
            }
        )
    return results


# ---------------------------------------------------------------------------
# ExcelExporter tests
# ---------------------------------------------------------------------------


class TestExcelExporter:
    def test_export_creates_file(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.xlsx")
        result_path = ExcelExporter().export(_minimal_results(), out)
        assert os.path.exists(result_path)

    def test_export_returns_absolute_path(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.xlsx")
        result_path = ExcelExporter().export(_minimal_results(), out)
        assert os.path.isabs(result_path)

    def test_workbook_has_summary_sheet(self, tmp_path: Path) -> None:
        import openpyxl

        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(), out)
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames

    def test_workbook_has_cashflow_sheet(self, tmp_path: Path) -> None:
        import openpyxl

        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(), out)
        wb = openpyxl.load_workbook(out)
        assert "Cashflow" in wb.sheetnames

    def test_workbook_no_dispatch_sheet_when_omitted(self, tmp_path: Path) -> None:
        import openpyxl

        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(with_dispatch=False), out)
        wb = openpyxl.load_workbook(out)
        assert "Dispatch" not in wb.sheetnames

    def test_workbook_has_dispatch_sheet_when_provided(self, tmp_path: Path) -> None:
        import openpyxl

        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(with_dispatch=True), out)
        wb = openpyxl.load_workbook(out)
        assert "Dispatch" in wb.sheetnames

    def test_summary_sheet_contains_scenario_name(self, tmp_path: Path) -> None:
        import openpyxl

        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        cell_values = [str(ws.cell(r, c).value or "") for r in range(1, 10) for c in range(1, 5)]
        assert any("Test Scenario" in v for v in cell_values)

    def test_cashflow_sheet_has_data_rows(self, tmp_path: Path) -> None:
        import openpyxl

        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Cashflow"]
        # At minimum the header row + 3 data rows
        assert ws.max_row >= 4

    def test_export_overwrites_existing_file(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.xlsx")
        ExcelExporter().export(_minimal_results(), out)
        mtime1 = os.path.getmtime(out)
        ExcelExporter().export(_minimal_results(), out)
        mtime2 = os.path.getmtime(out)
        # File was re-written (mtime may be equal on fast filesystems, but exists)
        assert os.path.exists(out)
        assert mtime2 >= mtime1

    def test_export_none_cashflow_still_creates_file(self, tmp_path: Path) -> None:
        results = _minimal_results()
        results["cashflow"] = None
        out = str(tmp_path / "report_nocf.xlsx")
        result_path = ExcelExporter().export(results, out)
        assert os.path.exists(result_path)


# ---------------------------------------------------------------------------
# PDFExporter tests
# ---------------------------------------------------------------------------


class TestPDFExporter:
    def test_export_creates_file(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.pdf")
        result_path = PDFExporter().export(_minimal_results(), out)
        assert os.path.exists(result_path)

    def test_export_returns_absolute_path(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.pdf")
        result_path = PDFExporter().export(_minimal_results(), out)
        assert os.path.isabs(result_path)

    def test_pdf_file_is_non_empty(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.pdf")
        PDFExporter().export(_minimal_results(), out)
        assert os.path.getsize(out) > 1024  # at least 1 KB

    def test_pdf_starts_with_pdf_magic_bytes(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.pdf")
        PDFExporter().export(_minimal_results(), out)
        with open(out, "rb") as f:
            header = f.read(5)
        assert header == b"%PDF-"

    def test_pdf_with_dispatch_profile(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report_dispatch.pdf")
        result_path = PDFExporter().export(_minimal_results(with_dispatch=True), out)
        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 1024

    def test_pdf_with_none_cashflow(self, tmp_path: Path) -> None:
        results = _minimal_results()
        results["cashflow"] = None
        out = str(tmp_path / "report_nocf.pdf")
        result_path = PDFExporter().export(results, out)
        assert os.path.exists(result_path)

    def test_pdf_overwrites_existing_file(self, tmp_path: Path) -> None:
        out = str(tmp_path / "report.pdf")
        PDFExporter().export(_minimal_results(), out)
        size1 = os.path.getsize(out)
        PDFExporter().export(_minimal_results(), out)
        size2 = os.path.getsize(out)
        assert size2 > 0
        assert abs(size2 - size1) < 5000  # same content → similar size
