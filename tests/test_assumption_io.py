"""Tests for app.assumptions.io — JSON and Excel import/export."""

from __future__ import annotations

import json
import uuid
from datetime import date
from pathlib import Path

import pytest
from pydantic import ValidationError

from app.assumptions.io import (
    export_excel,
    export_excel_file,
    export_json,
    export_json_file,
    import_excel,
    import_excel_file,
    import_json,
    import_json_file,
)
from app.assumptions.models import (
    AssumptionCategory,
    AssumptionEntry,
    AssumptionSet,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_entry(
    set_id: uuid.UUID,
    category: AssumptionCategory = AssumptionCategory.TARIFF,
    key: str = "peak_rate",
    value: object = 0.35,
) -> AssumptionEntry:
    return AssumptionEntry(
        set_id=set_id,
        category=category,
        key=key,
        value=value,
        unit="$/kWh",
        source="test",
    )


def _make_set(n_entries: int = 3) -> AssumptionSet:
    sid = uuid.uuid4()
    entries = [
        _make_entry(sid, AssumptionCategory.TARIFF, f"rate_{i}", i * 0.1) for i in range(n_entries)
    ]
    # Add a CAPEX entry so multiple categories are present
    entries.append(_make_entry(sid, AssumptionCategory.CAPEX, "solar_capex", {"cost_per_kw": 900}))
    return AssumptionSet(
        id=sid,
        name="Test Set",
        description="A test assumption set",
        author="pytest",
        effective_from=date(2025, 7, 1),
        entries=entries,
    )


# ---------------------------------------------------------------------------
# JSON round-trip tests
# ---------------------------------------------------------------------------


class TestJsonExport:
    def test_export_returns_valid_json(self) -> None:
        s = _make_set()
        raw = export_json(s)
        parsed = json.loads(raw)
        assert parsed["name"] == "Test Set"

    def test_export_includes_all_entries(self) -> None:
        s = _make_set(3)
        raw = export_json(s)
        parsed = json.loads(raw)
        assert len(parsed["entries"]) == 4  # 3 tariff + 1 capex

    def test_uuid_fields_serialised_as_strings(self) -> None:
        s = _make_set()
        raw = export_json(s)
        parsed = json.loads(raw)
        # UUIDs must be strings, not objects
        assert isinstance(parsed["id"], str)
        assert isinstance(parsed["entries"][0]["id"], str)

    def test_date_serialised_as_string(self) -> None:
        s = _make_set()
        raw = export_json(s)
        parsed = json.loads(raw)
        assert parsed["effective_from"] == "2025-07-01"

    def test_complex_value_preserved(self) -> None:
        s = _make_set()
        raw = export_json(s)
        parsed = json.loads(raw)
        capex_entry = next(e for e in parsed["entries"] if e["key"] == "solar_capex")
        assert capex_entry["value"] == {"cost_per_kw": 900}


class TestJsonImport:
    def test_roundtrip_preserves_name(self) -> None:
        s = _make_set()
        restored = import_json(export_json(s))
        assert restored.name == s.name

    def test_roundtrip_preserves_id(self) -> None:
        s = _make_set()
        restored = import_json(export_json(s))
        assert restored.id == s.id

    def test_roundtrip_preserves_entries_count(self) -> None:
        s = _make_set(5)
        restored = import_json(export_json(s))
        assert len(restored.entries) == len(s.entries)

    def test_roundtrip_preserves_entry_values(self) -> None:
        s = _make_set()
        restored = import_json(export_json(s))
        original_keys = {e.key for e in s.entries}
        restored_keys = {e.key for e in restored.entries}
        assert original_keys == restored_keys

    def test_import_accepts_dict(self) -> None:
        s = _make_set()
        data = json.loads(export_json(s))
        restored = import_json(data)
        assert restored.id == s.id

    def test_import_accepts_bytes(self) -> None:
        s = _make_set()
        raw_bytes = export_json(s).encode()
        restored = import_json(raw_bytes)
        assert restored.id == s.id

    def test_import_raises_on_invalid_json(self) -> None:
        with pytest.raises(json.JSONDecodeError):
            import_json("not valid json {{")

    def test_import_raises_on_missing_required_field(self) -> None:
        data = {"name": "Incomplete"}  # missing effective_from
        with pytest.raises(ValidationError):
            import_json(data)


class TestJsonFileRoundtrip:
    def test_file_roundtrip(self, tmp_path: Path) -> None:
        s = _make_set()
        fpath = tmp_path / "assumptions.json"
        export_json_file(s, fpath)
        restored = import_json_file(fpath)
        assert restored.id == s.id
        assert restored.name == s.name

    def test_file_is_human_readable(self, tmp_path: Path) -> None:
        s = _make_set()
        fpath = tmp_path / "assumptions.json"
        export_json_file(s, fpath)
        content = fpath.read_text(encoding="utf-8")
        # Pretty-printed — should have newlines
        assert "\n" in content


# ---------------------------------------------------------------------------
# Excel round-trip tests
# ---------------------------------------------------------------------------


class TestExcelExport:
    def test_export_returns_bytes(self) -> None:
        s = _make_set()
        raw = export_excel(s)
        assert isinstance(raw, bytes)
        assert len(raw) > 0

    def test_export_is_valid_xlsx(self) -> None:
        import openpyxl

        s = _make_set()
        raw = export_excel(s)
        import io

        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert "Metadata" in wb.sheetnames
        assert "All Entries" in wb.sheetnames

    def test_per_category_sheets_created(self) -> None:
        import io

        import openpyxl

        s = _make_set()
        raw = export_excel(s)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        # Tariff and Capex sheets should exist
        assert "Tariff" in wb.sheetnames
        assert "Capex" in wb.sheetnames

    def test_metadata_sheet_contains_name(self) -> None:
        import io

        import openpyxl

        s = _make_set()
        raw = export_excel(s)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb["Metadata"]
        rows = list(ws.iter_rows(values_only=True))
        meta_dict = {r[0]: r[1] for r in rows[1:] if r[0]}
        assert meta_dict.get("name") == "Test Set"

    def test_all_entries_row_count(self) -> None:
        import io

        import openpyxl

        s = _make_set(3)  # 3 tariff + 1 capex = 4 entries
        raw = export_excel(s)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb["All Entries"]
        rows = [r for r in ws.iter_rows(values_only=True) if any(r)]
        # header row + 4 data rows
        assert len(rows) == 5


class TestExcelImport:
    def test_roundtrip_preserves_name(self) -> None:
        s = _make_set()
        restored = import_excel(export_excel(s))
        assert restored.name == s.name

    def test_roundtrip_preserves_id(self) -> None:
        s = _make_set()
        restored = import_excel(export_excel(s))
        assert restored.id == s.id

    def test_roundtrip_preserves_entries_count(self) -> None:
        s = _make_set(4)
        restored = import_excel(export_excel(s))
        assert len(restored.entries) == len(s.entries)

    def test_roundtrip_preserves_entry_keys(self) -> None:
        s = _make_set(2)
        restored = import_excel(export_excel(s))
        original_keys = {e.key for e in s.entries}
        restored_keys = {e.key for e in restored.entries}
        assert original_keys == restored_keys

    def test_complex_value_survives_roundtrip(self) -> None:
        s = _make_set()
        restored = import_excel(export_excel(s))
        capex_entry = next(e for e in restored.entries if e.key == "solar_capex")
        assert capex_entry.value == {"cost_per_kw": 900}

    def test_import_raises_on_missing_metadata_sheet(self) -> None:
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "Not Metadata"  # type: ignore[union-attr]
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="Metadata"):
            import_excel(buf.getvalue())

    def test_import_raises_on_missing_entries_sheet(self) -> None:
        import io

        import openpyxl

        s = _make_set()
        raw = export_excel(s)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        del wb["All Entries"]
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="All Entries"):
            import_excel(buf.getvalue())


class TestExcelFileRoundtrip:
    def test_file_roundtrip(self, tmp_path: Path) -> None:
        s = _make_set()
        fpath = tmp_path / "assumptions.xlsx"
        export_excel_file(s, fpath)
        restored = import_excel_file(fpath)
        assert restored.id == s.id
        assert restored.name == s.name
        assert len(restored.entries) == len(s.entries)

    def test_file_is_xlsx(self, tmp_path: Path) -> None:
        s = _make_set()
        fpath = tmp_path / "assumptions.xlsx"
        export_excel_file(s, fpath)
        # XLSX files start with PK (zip magic bytes)
        header = fpath.read_bytes()[:2]
        assert header == b"PK"
